# excel_exporter.py
from PyQt5.QtCore import QRunnable, pyqtSignal, QObject
from openpyxl.drawing.image import Image
from PIL import Image as PilImage
import openpyxl

class ExportSignals(QObject):
    progress = pyqtSignal(int)
    finished = pyqtSignal()
    error = pyqtSignal(str)

class ExcelExporter(QRunnable):
    def __init__(self, db_manager, filtered_suppliers, selected_file, include_images):
        super().__init__()
        self.db_manager = db_manager
        self.filtered_suppliers = filtered_suppliers
        self.selected_file = selected_file
        self.include_images = include_images
        self.signals = ExportSignals()

    def run(self):
        try:
            # 定义图片目录路径变量
            image_directory = "//VIVA303-WORK/Viva店面共享/StockImg/"

            # 获取要导出的数据（从数据库中获得）
            selected_suppliers = self.filtered_suppliers
            rows_to_export = self.db_manager.fetch_rugs()

            # 指定自定义的列顺序
            column_order = ['图片', '型号', '类型', '数量', '供货商', '备注']

            # 对数据进行排序以符合要求
            rows_to_export.sort(key=lambda x: (x[2], x[1], x[0]))

            # 创建一个新的 Excel 工作簿
            workbook = openpyxl.Workbook()

            # 创建一个默认的工作表
            default_sheet = workbook.active
            default_sheet.title = '所有供货商'

            # 添加列标题
            default_sheet.append(column_order)

            # 设置列宽和行高，如果需要包含图片
            if self.include_images:
                default_sheet.column_dimensions['A'].width = 15
                default_sheet.column_dimensions['B'].width = 30
                default_sheet.column_dimensions['F'].width = 30
                for row_index in range(2, len(rows_to_export) + 2):
                    default_sheet.row_dimensions[row_index].height = 80

            # 将数据填充到默认工作表
            for row in rows_to_export:
                # 提取按照自定义顺序的列数据
                row_data = [row[5] if not self.include_images else None, row[0], row[3], row[1], row[2], row[4]]
                default_sheet.append(row_data)
                if self.include_images:
                    # 插入图片
                    img_path = f"{image_directory}{row[5]}"
                    try:
                        with PilImage.open(img_path) as img:
                            # 获取单元格宽高
                            cell_width = default_sheet.column_dimensions['A'].width * 7.2
                            cell_height = default_sheet.row_dimensions[default_sheet.max_row].height

                            # 计算图片的缩放比例，保持图片比例并适应单元格
                            width_ratio = cell_width / img.width
                            height_ratio = cell_height / img.height
                            scale_ratio = min(width_ratio, height_ratio)
                            new_width = int(img.width * scale_ratio)
                            new_height = int(img.height * scale_ratio)
                            img = img.resize((new_width, new_height), PilImage.LANCZOS)
                            
                            temp_path = f"temp_{row[5]}"
                            img.save(temp_path)
                            excel_img = Image(temp_path)
                            # 插入图片
                            img_col = 'A'
                            img_row = default_sheet.max_row
                            img_cell = default_sheet.cell(row=img_row, column=1)
                            img_cell.value = None
                            default_sheet.add_image(excel_img, f'{img_col}{img_row}')
                    except Exception:
                        pass

            # 获取不同供货商的数据
            suppliers = set(row[2] for row in rows_to_export)

            # 创建一个工作表来存储每个供货商的数据
            for supplier in suppliers:
                sheet = workbook.create_sheet(title=supplier)

                # 添加列标题
                sheet.append(column_order)

                # 设置列宽和行高，如果需要包含图片
                if self.include_images:
                    sheet.column_dimensions['A'].width = 15
                    sheet.column_dimensions['B'].width = 30
                    sheet.column_dimensions['F'].width = 30
                    for row_index in range(2, len(rows_to_export) + 2):
                        sheet.row_dimensions[row_index].height = 80

                supplier_data = [row for row in rows_to_export if row[2] == supplier]
                for row in supplier_data:
                    # 按照指定顺序提取数据列
                    data_columns = [row[5] if not self.include_images else None, row[0], row[3], row[1], row[2], row[4]]
                    sheet.append(data_columns)
                    if self.include_images:
                        # 插入图片
                        img_path = f"{image_directory}{row[5]}"
                        try:
                            with PilImage.open(img_path) as img:
                                # 获取单元格宽高
                                cell_width = sheet.column_dimensions['A'].width * 7.2
                                cell_height = sheet.row_dimensions[sheet.max_row].height

                                # 计算图片的缩放比例，保持图片比例并适应单元格
                                width_ratio = cell_width / img.width
                                height_ratio = cell_height / img.height
                                scale_ratio = min(width_ratio, height_ratio)
                                new_width = int(img.width * scale_ratio)
                                new_height = int(img.height * scale_ratio)
                                img = img.resize((new_width, new_height), PilImage.LANCZOS)
                                
                                temp_path = f"temp_{row[5]}"
                                img.save(temp_path)
                                excel_img = Image(temp_path)
                                # 插入图片
                                img_col = 'A'
                                img_row = sheet.max_row
                                img_cell = sheet.cell(row=img_row, column=1)
                                img_cell.value = None
                                sheet.add_image(excel_img, f'{img_col}{img_row}')
                        except Exception:
                            pass

            # 保存 Excel 文件
            workbook.save(self.selected_file)

            # 发送完成信号
            self.signals.finished.emit()
        except Exception as e:
            self.signals.error.emit(str(e))
